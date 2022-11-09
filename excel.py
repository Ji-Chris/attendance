# import xlsxwriter
import openpyxl
import datetime
import os

SAVE_PATH = "attendance.xlsx"
WORKSHEET_NAME = "Automatic Attendance"

wb = None
ws = None

PRESENT_STRING = "P" 
ABSENT_STRING = "A"

def save_attendance(attendance : set[str]):
    
    global wb
    global ws
    
    # Read the excel file
    wb = openpyxl.load_workbook(SAVE_PATH)
    ws = wb[WORKSHEET_NAME]

    # get latest date
    column_n = 1
    today = datetime.datetime.today().strftime("%d-%m-%Y")
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for cell in row:
            if cell == today:
                break
            column_n += 1
    else:
        ws.cell(1, column_n, today)
    
    row_n = 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        row_n+=1
        for cell in row:
            if cell in attendance:
                ws.cell(row=row_n, column=column_n).value = PRESENT_STRING
            else:
                ws.cell(row=row_n, column=column_n).value = ABSENT_STRING
    
    # Save workbook
    wb.save(SAVE_PATH)

def setup(students : list[str]):
    
    wb = openpyxl.Workbook(SAVE_PATH)
    ws = wb.create_sheet(WORKSHEET_NAME)

    # write header row
    ws.append(["Name"])
    for i, student in enumerate(students):
        ws.append([student])
    
    wb.save(SAVE_PATH)

    